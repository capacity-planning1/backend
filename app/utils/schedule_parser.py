import io

import openpyxl
import pandas as pd
import requests
from sqlalchemy import insert
from sqlmodel import delete
from sqlmodel.ext.asyncio.session import AsyncSession

from app.models.schedule import ScheduleModel, WeekDay


def parse_schedule_to_dataframe(sheet_id):
    url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx'

    print('загрузка')
    response = requests.get(url)
    if response.status_code != 200:
        print('не удалось загрузить')
        return None

    wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
    sheet = wb.active

    groups = {}
    for col in range(3, sheet.max_column + 1):
        group_ceil = sheet.cell(row=2, column=col).value
        if group_ceil:
            groups[col] = str(group_ceil).strip()

    merged_ranges = sheet.merged_cells.ranges

    def get_cell_value_styled(row, col):
        for merged_range in merged_ranges:
            if row in range(
                merged_range.min_row, merged_range.max_row + 1
            ) and col in range(merged_range.min_col, merged_range.max_col + 1):
                return sheet.cell(
                    row=merged_range.min_row, column=merged_range.min_col
                ).value
        return sheet.cell(row=row, column=col).value

    schedule_data = []

    for row in range(3, sheet.max_row + 1):
        raw_day = get_cell_value_styled(row, 1)
        time_slot = get_cell_value_styled(row, 2)

        if not raw_day or not time_slot:
            continue

        clean_day_str = str(raw_day).strip().replace("\n", "").capitalize()

        try:
            day_enum = WeekDay(clean_day_str)
        except ValueError:
            continue

        for col, group_name in groups.items():
            cell_content = get_cell_value_styled(row, col)

            if cell_content:
                lesson_text = " ".join(str(cell_content).split())

                schedule_data.append(
                    {
                        "day": day_enum.value, 
                        "time_slot": str(time_slot).strip(),
                        "group": group_name,
                        "lesson_details": lesson_text,
                    }
                )

    return pd.DataFrame(schedule_data)


async def export_to_postgres(df: pd.DataFrame, session: AsyncSession):
    if df is None or df.empty:
        return

    records = df.to_dict(orient="records")
    if records:
        await session.exec(delete(ScheduleModel))
        await session.execute(insert(ScheduleModel), records)

    await session.commit()
